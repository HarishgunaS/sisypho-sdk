"""
Core library for Excel file manipulation using openpyxl.
"""

from typing import Dict, Any, Optional, List, Union, Tuple
import logging
from pathlib import Path
import os as std_os

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
except ImportError as e:
    logger.error(f"openpyxl not installed. Install with: pip install openpyxl")
    raise ImportError("openpyxl library is required for Excel operations") from e

# Global workbook cache for performance
_workbook_cache: Dict[str, Workbook] = {}

def _get_workbook(file_path: str, create_if_missing: bool = False) -> Optional[Workbook]:
    """
    Get workbook from cache or load from file.
    
    Args:
        file_path: Path to the Excel file.
        create_if_missing: Create new workbook if file doesn't exist.
        
    Returns:
        Workbook instance or None if failed.
    """
    try:
        abs_path = std_os.path.abspath(file_path)
        
        # Check cache first
        if abs_path in _workbook_cache:
            return _workbook_cache[abs_path]
            
        # Load from file or create new
        if std_os.path.exists(file_path):
            wb = load_workbook(file_path)
        elif create_if_missing:
            wb = Workbook()
        else:
            logger.error(f"File not found: {file_path}")
            return None
            
        _workbook_cache[abs_path] = wb
        return wb
    except Exception as e:
        logger.error(f"Failed to get workbook '{file_path}': {e}")
        return None

def _clear_cache(file_path: Optional[str] = None):
    """Clear workbook cache for specific file or all files."""
    global _workbook_cache
    if file_path:
        abs_path = std_os.path.abspath(file_path)
        _workbook_cache.pop(abs_path, None)
    else:
        _workbook_cache.clear()

# File Operations
def create_workbook(file_path: str, overwrite: bool = False) -> bool:
    """
    Create a new Excel workbook.
    
    Args:
        file_path: Path where the workbook will be saved.
        overwrite: Whether to overwrite existing file.
        
    Returns:
        True if workbook was created successfully, False otherwise.
        
    Examples:
        >>> create_workbook("new_report.xlsx")
        True
        >>> create_workbook("backup.xlsx", overwrite=True)
        True
    """
    try:
        if std_os.path.exists(file_path) and not overwrite:
            logger.error(f"File already exists: {file_path}")
            return False
            
        # Create directory if it doesn't exist
        std_os.makedirs(std_os.path.dirname(std_os.path.abspath(file_path)), exist_ok=True)
        
        wb = Workbook()
        wb.save(file_path)
        
        # Cache the workbook
        _workbook_cache[std_os.path.abspath(file_path)] = wb
        
        logger.info(f"Created workbook: {file_path}")
        return True
    except Exception as e:
        logger.error(f"Failed to create workbook '{file_path}': {e}")
        return False

def open_workbook(file_path: str) -> bool:
    """
    Open an existing Excel workbook.
    
    Args:
        file_path: Path to the Excel file.
        
    Returns:
        True if workbook was opened successfully, False otherwise.
        
    Examples:
        >>> open_workbook("data.xlsx")
        True
        >>> open_workbook("reports/monthly.xlsx")
        True
    """
    try:
        wb = _get_workbook(file_path, create_if_missing=False)
        return wb is not None
    except Exception as e:
        logger.error(f"Failed to open workbook '{file_path}': {e}")
        return False

def save_workbook(file_path: str, save_as_path: Optional[str] = None) -> bool:
    """
    Save a workbook to file.
    
    Args:
        file_path: Path to the current workbook.
        save_as_path: Optional path to save as a different file.
        
    Returns:
        True if workbook was saved successfully, False otherwise.
        
    Examples:
        >>> save_workbook("data.xlsx")
        True
        >>> save_workbook("data.xlsx", "backup/data_backup.xlsx")
        True
    """
    try:
        wb = _get_workbook(file_path)
        if not wb:
            return False
            
        target_path = save_as_path or file_path
        
        # Create directory if it doesn't exist
        std_os.makedirs(std_os.path.dirname(std_os.path.abspath(target_path)), exist_ok=True)
        
        wb.save(target_path)
        
        # Update cache if saved to different location
        if save_as_path:
            _workbook_cache[std_os.path.abspath(save_as_path)] = wb
            
        logger.info(f"Saved workbook to: {target_path}")
        return True
    except Exception as e:
        logger.error(f"Failed to save workbook '{file_path}': {e}")
        return False

def close_workbook(file_path: str) -> bool:
    """
    Close a workbook and remove from cache.
    
    Args:
        file_path: Path to the workbook to close.
        
    Returns:
        True if workbook was closed successfully, False otherwise.
        
    Examples:
        >>> close_workbook("data.xlsx")
        True
    """
    try:
        _clear_cache(file_path)
        logger.info(f"Closed workbook: {file_path}")
        return True
    except Exception as e:
        logger.error(f"Failed to close workbook '{file_path}': {e}")
        return False

# Worksheet Operations
def create_worksheet(file_path: str, sheet_name: str, index: Optional[int] = None) -> bool:
    """
    Create a new worksheet in the workbook.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name for the new worksheet.
        index: Optional position index for the worksheet.
        
    Returns:
        True if worksheet was created successfully, False otherwise.
        
    Examples:
        >>> create_worksheet("data.xlsx", "Summary")
        True
        >>> create_worksheet("data.xlsx", "Q1 Data", 0)
        True
    """
    try:
        wb = _get_workbook(file_path, create_if_missing=True)
        if not wb:
            return False
            
        if index is not None:
            ws = wb.create_sheet(sheet_name, index)
        else:
            ws = wb.create_sheet(sheet_name)
            
        logger.info(f"Created worksheet '{sheet_name}' in {file_path}")
        wb.save(file_path)
        return True
    except Exception as e:
        logger.error(f"Failed to create worksheet '{sheet_name}': {e}")
        return False

def delete_worksheet(file_path: str, sheet_name: str) -> bool:
    """
    Delete a worksheet from the workbook.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet to delete.
        
    Returns:
        True if worksheet was deleted successfully, False otherwise.
        
    Examples:
        >>> delete_worksheet("data.xlsx", "Temp")
        True
    """
    try:
        wb = _get_workbook(file_path)
        if not wb:
            return False
            
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
            logger.info(f"Deleted worksheet '{sheet_name}' from {file_path}")
            wb.save(file_path)
            return True
        else:
            logger.error(f"Worksheet '{sheet_name}' not found")
            return False
    except Exception as e:
        logger.error(f"Failed to delete worksheet '{sheet_name}': {e}")
        return False

def rename_worksheet(file_path: str, old_name: str, new_name: str) -> bool:
    """
    Rename a worksheet.
    
    Args:
        file_path: Path to the Excel file.
        old_name: Current name of the worksheet.
        new_name: New name for the worksheet.
        
    Returns:
        True if worksheet was renamed successfully, False otherwise.
        
    Examples:
        >>> rename_worksheet("data.xlsx", "Sheet1", "Data")
        True
    """
    try:
        wb = _get_workbook(file_path)
        if not wb:
            return False
            
        if old_name in wb.sheetnames:
            wb[old_name].title = new_name
            logger.info(f"Renamed worksheet '{old_name}' to '{new_name}'")
            wb.save(file_path)
            return True
        else:
            logger.error(f"Worksheet '{old_name}' not found")
            return False
    except Exception as e:
        logger.error(f"Failed to rename worksheet '{old_name}': {e}")
        return False

def list_worksheets(file_path: str) -> List[str]:
    """
    Get list of all worksheet names in the workbook.
    
    Args:
        file_path: Path to the Excel file.
        
    Returns:
        List of worksheet names, empty list if failed.
        
    Examples:
        >>> sheets = list_worksheets("data.xlsx")
        >>> print(sheets)
        ['Sheet1', 'Data', 'Summary']
    """
    try:
        wb = _get_workbook(file_path)
        if not wb:
            return []
        return wb.sheetnames
    except Exception as e:
        logger.error(f"Failed to list worksheets in '{file_path}': {e}")
        return []

# Cell Operations
def read_cell(file_path: str, sheet_name: str, cell: str) -> Any:
    """
    Read value from a specific cell.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet.
        cell: Cell reference (e.g., 'A1', 'B5').
        
    Returns:
        Cell value or None if failed.
        
    Examples:
        >>> value = read_cell("data.xlsx", "Sheet1", "A1")
        >>> print(value)
        'Hello World'
        >>> number = read_cell("data.xlsx", "Sheet1", "B2")
        >>> print(number)
        42
    """
    try:
        wb = _get_workbook(file_path)
        if not wb or sheet_name not in wb.sheetnames:
            return None
            
        ws = wb[sheet_name]
        return ws[cell].value
    except Exception as e:
        logger.error(f"Failed to read cell '{cell}' from '{sheet_name}': {e}")
        return None

def write_cell(file_path: str, sheet_name: str, cell: str, value: Any) -> bool:
    """
    Write value to a specific cell.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet.
        cell: Cell reference (e.g., 'A1', 'B5').
        value: Value to write to the cell.
        
    Returns:
        True if cell was written successfully, False otherwise.
        
    Examples:
        >>> write_cell("data.xlsx", "Sheet1", "A1", "Hello World")
        True
        >>> write_cell("data.xlsx", "Sheet1", "B2", 42)
        True
        >>> write_cell("data.xlsx", "Sheet1", "C3", "=A1+B2")
        True
    """
    try:
        wb = _get_workbook(file_path, create_if_missing=True)
        if not wb:
            return False
            
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
            
        ws = wb[sheet_name]
        ws[cell] = value
        
        logger.info(f"Wrote value to cell '{cell}' in '{sheet_name}'")
        wb.save(file_path)
        return True
    except Exception as e:
        logger.error(f"Failed to write cell '{cell}': {e}")
        return False

def read_range(file_path: str, sheet_name: str, start_cell: str, end_cell: str) -> List[List[Any]]:
    """
    Read values from a range of cells.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet.
        start_cell: Starting cell reference (e.g., 'A1').
        end_cell: Ending cell reference (e.g., 'C5').
        
    Returns:
        2D list of cell values, empty list if failed.
        
    Examples:
        >>> data = read_range("data.xlsx", "Sheet1", "A1", "C3")
        >>> print(data)
        [['Name', 'Age', 'City'], ['John', 25, 'NYC'], ['Jane', 30, 'LA']]
    """
    try:
        wb = _get_workbook(file_path)
        if not wb or sheet_name not in wb.sheetnames:
            return []
            
        ws = wb[sheet_name]
        cell_range = f"{start_cell}:{end_cell}"
        
        result = []
        for row in ws[cell_range]:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            result.append(row_data)
            
        return result
    except Exception as e:
        logger.error(f"Failed to read range '{start_cell}:{end_cell}': {e}")
        return []

def write_range(file_path: str, sheet_name: str, start_cell: str, data: List[List[Any]]) -> bool:
    """
    Write data to a range of cells starting from the specified cell.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet.
        start_cell: Starting cell reference (e.g., 'A1').
        data: 2D list of values to write.
        
    Returns:
        True if range was written successfully, False otherwise.
        
    Examples:
        >>> data = [['Name', 'Age', 'City'], ['John', 25, 'NYC'], ['Jane', 30, 'LA']]
        >>> write_range("data.xlsx", "Sheet1", "A1", data)
        True
    """
    try:
        wb = _get_workbook(file_path, create_if_missing=True)
        if not wb:
            return False
            
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
            
        ws = wb[sheet_name]
        
        # Parse start cell to get row and column indices
        start_col = column_index_from_string(start_cell.rstrip('0123456789'))
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        
        # Write data
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)
        
        wb.save(file_path)
                
        logger.info(f"Wrote {len(data)}x{len(data[0]) if data else 0} range starting at '{start_cell}'")
        return True
    except Exception as e:
        logger.error(f"Failed to write range starting at '{start_cell}': {e}")
        return False

def clear_range(file_path: str, sheet_name: str, start_cell: str, end_cell: str) -> bool:
    """
    Clear values from a range of cells.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet.
        start_cell: Starting cell reference (e.g., 'A1').
        end_cell: Ending cell reference (e.g., 'C5').
        
    Returns:
        True if range was cleared successfully, False otherwise.
        
    Examples:
        >>> clear_range("data.xlsx", "Sheet1", "A1", "C10")
        True
    """
    try:
        wb = _get_workbook(file_path)
        if not wb or sheet_name not in wb.sheetnames:
            return False
            
        ws = wb[sheet_name]
        cell_range = f"{start_cell}:{end_cell}"
        
        for row in ws[cell_range]:
            for cell in row:
                cell.value = None
                
        wb.save(file_path)
        logger.info(f"Cleared range '{start_cell}:{end_cell}' in '{sheet_name}'")
        return True
    except Exception as e:
        logger.error(f"Failed to clear range '{start_cell}:{end_cell}': {e}")
        return False

# Row and Column Operations
def insert_rows(file_path: str, sheet_name: str, row_index: int, count: int = 1) -> bool:
    """
    Insert rows at the specified index.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet.
        row_index: Row index where to insert (1-based).
        count: Number of rows to insert.
        
    Returns:
        True if rows were inserted successfully, False otherwise.
        
    Examples:
        >>> insert_rows("data.xlsx", "Sheet1", 3, 2)
        True
    """
    try:
        wb = _get_workbook(file_path)
        if not wb or sheet_name not in wb.sheetnames:
            return False
            
        ws = wb[sheet_name]
        ws.insert_rows(row_index, count)
        
        logger.info(f"Inserted {count} rows at index {row_index}")
        wb.save(file_path)
        return True
    except Exception as e:
        logger.error(f"Failed to insert rows: {e}")
        return False

def insert_columns(file_path: str, sheet_name: str, col_index: int, count: int = 1) -> bool:
    """
    Insert columns at the specified index.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet.
        col_index: Column index where to insert (1-based).
        count: Number of columns to insert.
        
    Returns:
        True if columns were inserted successfully, False otherwise.
        
    Examples:
        >>> insert_columns("data.xlsx", "Sheet1", 2, 1)
        True
    """
    try:
        wb = _get_workbook(file_path)
        if not wb or sheet_name not in wb.sheetnames:
            return False
            
        ws = wb[sheet_name]
        ws.insert_cols(col_index, count)
        
        logger.info(f"Inserted {count} columns at index {col_index}")
        wb.save(file_path)
        return True
    except Exception as e:
        logger.error(f"Failed to insert columns: {e}")
        return False

def delete_rows(file_path: str, sheet_name: str, row_index: int, count: int = 1) -> bool:
    """
    Delete rows at the specified index.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet.
        row_index: Row index to start deleting from (1-based).
        count: Number of rows to delete.
        
    Returns:
        True if rows were deleted successfully, False otherwise.
        
    Examples:
        >>> delete_rows("data.xlsx", "Sheet1", 5, 3)
        True
    """
    try:
        wb = _get_workbook(file_path)
        if not wb or sheet_name not in wb.sheetnames:
            return False
            
        ws = wb[sheet_name]
        ws.delete_rows(row_index, count)
        
        logger.info(f"Deleted {count} rows starting at index {row_index}")
        wb.save(file_path)
        return True
    except Exception as e:
        logger.error(f"Failed to delete rows: {e}")
        return False

def delete_columns(file_path: str, sheet_name: str, col_index: int, count: int = 1) -> bool:
    """
    Delete columns at the specified index.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet.
        col_index: Column index to start deleting from (1-based).
        count: Number of columns to delete.
        
    Returns:
        True if columns were deleted successfully, False otherwise.
        
    Examples:
        >>> delete_columns("data.xlsx", "Sheet1", 3, 2)
        True
    """
    try:
        wb = _get_workbook(file_path)
        if not wb or sheet_name not in wb.sheetnames:
            return False
            
        ws = wb[sheet_name]
        ws.delete_cols(col_index, count)
        
        logger.info(f"Deleted {count} columns starting at index {col_index}")
        wb.save(file_path)
        return True
    except Exception as e:
        logger.error(f"Failed to delete columns: {e}")
        return False

# Formatting Operations
def format_cells(file_path: str, sheet_name: str, cell_range: str,
                font_name: Optional[str] = None, font_size: Optional[int] = None,
                bold: Optional[bool] = None, italic: Optional[bool] = None,
                font_color: Optional[str] = None, bg_color: Optional[str] = None,
                alignment: Optional[str] = None) -> bool:
    """
    Apply formatting to a range of cells.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet.
        cell_range: Cell range to format (e.g., 'A1:C5').
        font_name: Font name (e.g., 'Arial').
        font_size: Font size in points.
        bold: Make text bold.
        italic: Make text italic.
        font_color: Font color in hex format (e.g., 'FF0000' for red).
        bg_color: Background color in hex format.
        alignment: Text alignment ('left', 'center', 'right').
        
    Returns:
        True if formatting was applied successfully, False otherwise.
        
    Examples:
        >>> format_cells("data.xlsx", "Sheet1", "A1:A10", bold=True, font_size=12)
        True
        >>> format_cells("data.xlsx", "Sheet1", "B1:B1", bg_color="FFFF00", alignment="center")
        True
    """
    try:
        wb = _get_workbook(file_path)
        if not wb or sheet_name not in wb.sheetnames:
            return False
            
        ws = wb[sheet_name]
        
        # Create font style
        font_kwargs = {}
        if font_name:
            font_kwargs['name'] = font_name
        if font_size:
            font_kwargs['size'] = font_size
        if bold is not None:
            font_kwargs['bold'] = bold
        if italic is not None:
            font_kwargs['italic'] = italic
        if font_color:
            font_kwargs['color'] = font_color
            
        font_style = Font(**font_kwargs) if font_kwargs else None
        
        # Create fill style
        fill_style = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid") if bg_color else None
        
        # Create alignment style
        alignment_style = None
        if alignment:
            alignment_map = {'left': 'left', 'center': 'center', 'right': 'right'}
            if alignment in alignment_map:
                alignment_style = Alignment(horizontal=alignment_map[alignment])
        
        # Apply formatting to range
        for row in ws[cell_range]:
            for cell in row:
                if font_style:
                    cell.font = font_style
                if fill_style:
                    cell.fill = fill_style
                if alignment_style:
                    cell.alignment = alignment_style
                    
        logger.info(f"Applied formatting to range '{cell_range}' in '{sheet_name}'")
        wb.save(file_path)
        return True
    except Exception as e:
        logger.error(f"Failed to format cells '{cell_range}': {e}")
        return False

def auto_fit_columns(file_path: str, sheet_name: str) -> bool:
    """
    Auto-fit column widths based on content.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet.
        
    Returns:
        True if columns were auto-fitted successfully, False otherwise.
        
    Examples:
        >>> auto_fit_columns("data.xlsx", "Sheet1")
        True
    """
    try:
        wb = _get_workbook(file_path)
        if not wb or sheet_name not in wb.sheetnames:
            return False
            
        ws = wb[sheet_name]
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                    
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
            
        logger.info(f"Auto-fitted columns in '{sheet_name}'")
        wb.save(file_path)
        return True
    except Exception as e:
        logger.error(f"Failed to auto-fit columns: {e}")
        return False

# Advanced Operations
def add_formula(file_path: str, sheet_name: str, cell: str, formula: str) -> bool:
    """
    Add a formula to a cell.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet.
        cell: Cell reference where to add the formula.
        formula: Excel formula (should start with '=').
        
    Returns:
        True if formula was added successfully, False otherwise.
        
    Examples:
        >>> add_formula("data.xlsx", "Sheet1", "C1", "=A1+B1")
        True
        >>> add_formula("data.xlsx", "Sheet1", "D5", "=SUM(A1:A10)")
        True
    """
    try:
        if not formula.startswith('='):
            formula = '=' + formula
            
        return write_cell(file_path, sheet_name, cell, formula)
    except Exception as e:
        logger.error(f"Failed to add formula '{formula}': {e}")
        return False

def find_and_replace(file_path: str, sheet_name: str, find_text: str, replace_text: str) -> int:
    """
    Find and replace text in a worksheet.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet.
        find_text: Text to find.
        replace_text: Text to replace with.
        
    Returns:
        Number of replacements made, -1 if failed.
        
    Examples:
        >>> count = find_and_replace("data.xlsx", "Sheet1", "old_value", "new_value")
        >>> print(f"Replaced {count} instances")
        Replaced 5 instances
    """
    try:
        wb = _get_workbook(file_path)
        if not wb or sheet_name not in wb.sheetnames:
            return -1
            
        ws = wb[sheet_name]
        replacements = 0
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value) == find_text:
                    cell.value = replace_text
                    replacements += 1
                    
        wb.save(file_path)
        logger.info(f"Made {replacements} replacements in '{sheet_name}'")
        return replacements
    except Exception as e:
        logger.error(f"Failed to find and replace: {e}")
        return -1

def merge_cells(file_path: str, sheet_name: str, start_cell: str, end_cell: str) -> bool:
    """
    Merge a range of cells.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet.
        start_cell: Starting cell reference (e.g., 'A1').
        end_cell: Ending cell reference (e.g., 'C3').
        
    Returns:
        True if cells were merged successfully, False otherwise.
        
    Examples:
        >>> merge_cells("data.xlsx", "Sheet1", "A1", "C1")
        True
    """
    try:
        wb = _get_workbook(file_path)
        if not wb or sheet_name not in wb.sheetnames:
            return False
            
        ws = wb[sheet_name]
        ws.merge_cells(f"{start_cell}:{end_cell}")
        
        logger.info(f"Merged cells '{start_cell}:{end_cell}' in '{sheet_name}'")
        wb.save(file_path)
        return True
    except Exception as e:
        logger.error(f"Failed to merge cells '{start_cell}:{end_cell}': {e}")
        return False

def unmerge_cells(file_path: str, sheet_name: str, start_cell: str, end_cell: str) -> bool:
    """
    Unmerge a range of cells.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet.
        start_cell: Starting cell reference (e.g., 'A1').
        end_cell: Ending cell reference (e.g., 'C3').
        
    Returns:
        True if cells were unmerged successfully, False otherwise.
        
    Examples:
        >>> unmerge_cells("data.xlsx", "Sheet1", "A1", "C1")
        True
    """
    try:
        wb = _get_workbook(file_path)
        if not wb or sheet_name not in wb.sheetnames:
            return False
            
        ws = wb[sheet_name]
        ws.unmerge_cells(f"{start_cell}:{end_cell}")
        
        logger.info(f"Unmerged cells '{start_cell}:{end_cell}' in '{sheet_name}'")
        wb.save(file_path)
        return True
    except Exception as e:
        logger.error(f"Failed to unmerge cells '{start_cell}:{end_cell}': {e}")
        return False

# Convenience Functions
def create_table(file_path: str, sheet_name: str, start_cell: str, 
                headers: List[str], data: List[List[Any]]) -> bool:
    """
    Create a formatted table with headers.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet.
        start_cell: Starting cell reference for the table.
        headers: List of column headers.
        data: 2D list of table data.
        
    Returns:
        True if table was created successfully, False otherwise.
        
    Examples:
        >>> headers = ['Name', 'Age', 'City']
        >>> data = [['John', 25, 'NYC'], ['Jane', 30, 'LA']]
        >>> create_table("data.xlsx", "Sheet1", "A1", headers, data)
        True
    """
    try:
        # Write headers and data
        table_data = [headers] + data
        if not write_range(file_path, sheet_name, start_cell, table_data):
            return False
            
        # Format headers
        wb = _get_workbook(file_path)
        if not wb or sheet_name not in wb.sheetnames:
            return False
            
        ws = wb[sheet_name]
        start_col = column_index_from_string(start_cell.rstrip('0123456789'))
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        end_col_letter = get_column_letter(start_col + len(headers) - 1)
        
        header_range = f"{start_cell}:{end_col_letter}{start_row}"
        
        # Apply header formatting
        format_cells(file_path, sheet_name, header_range, 
                    bold=True, bg_color="D3D3D3", alignment="center")
        
        # Auto-fit columns
        auto_fit_columns(file_path, sheet_name)
        
        wb.save(file_path)
        logger.info(f"Created table with {len(headers)} columns and {len(data)} rows")
        return True
    except Exception as e:
        logger.error(f"Failed to create table: {e}")
        return False

def append_row_to_table(file_path: str, sheet_name: str, row_data: List[Any], 
                       table_start_row: int = 1) -> bool:
    """
    Append a row to the end of an existing table.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet.
        row_data: List of values for the new row.
        table_start_row: Row where the table starts (1-based).
        
    Returns:
        True if row was appended successfully, False otherwise.
        
    Examples:
        >>> append_row_to_table("data.xlsx", "Sheet1", ["Bob", 35, "Chicago"])
        True
    """
    try:
        wb = _get_workbook(file_path)
        if not wb or sheet_name not in wb.sheetnames:
            return False
            
        ws = wb[sheet_name]
        
        # Find the last row with data
        last_row = table_start_row
        while ws[f"A{last_row}"].value is not None:
            last_row += 1
            
        # Write the new row
        start_cell = f"A{last_row}"
        return write_range(file_path, sheet_name, start_cell, [row_data])
    except Exception as e:
        logger.error(f"Failed to append row to table: {e}")
        return False

def get_table_data(file_path: str, sheet_name: str, start_cell: str = "A1") -> Dict[str, Any]:
    """
    Read table data including headers and rows.
    
    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the worksheet.
        start_cell: Starting cell of the table.
        
    Returns:
        Dict with 'headers' and 'data' keys, empty dict if failed.
        
    Examples:
        >>> table = get_table_data("data.xlsx", "Sheet1")
        >>> print(table['headers'])
        ['Name', 'Age', 'City']
        >>> print(len(table['data']))
        25
    """
    try:
        wb = _get_workbook(file_path)
        if not wb or sheet_name not in wb.sheetnames:
            return {}
            
        ws = wb[sheet_name]
        
        # Parse start cell
        start_col = column_index_from_string(start_cell.rstrip('0123456789'))
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        
        # Find table dimensions
        max_col = start_col
        max_row = start_row
        
        # Find last column with data in header row
        while ws.cell(row=start_row, column=max_col).value is not None:
            max_col += 1
        max_col -= 1
        
        # Find last row with data
        while ws.cell(row=max_row, column=start_col).value is not None:
            max_row += 1
        max_row -= 1
        
        if max_col < start_col or max_row < start_row:
            return {}
            
        # Read data
        end_cell = f"{get_column_letter(max_col)}{max_row}"
        all_data = read_range(file_path, sheet_name, start_cell, end_cell)
        
        if not all_data:
            return {}
            
        return {
            'headers': all_data[0] if all_data else [],
            'data': all_data[1:] if len(all_data) > 1 else []
        }
    except Exception as e:
        logger.error(f"Failed to get table data: {e}")
        return {}